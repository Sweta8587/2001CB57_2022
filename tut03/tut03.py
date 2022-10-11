from math import ceil
import openpyxl
from openpyxl import Workbook
from datetime import datetime

start_time = datetime.now()

def get_octant(x, y, z):
    if(x>=0 and y>=0 and z>=0):
        return(1)
    elif(x>=0 and y>=0 and z<0):
        return(-1)
    elif(x<0 and y>=0 and z>=0):
        return(2)
    elif(x<0 and y>=0 and z<0):
        return(-2)
    elif(x<0 and y<0 and z>=0):
        return(3)
    elif(x<0 and y<0 and z<0):
        return(-3)
    elif(x>=0 and y<0 and z>=0):
        return(4)
    else:
        return(-4)

def octant_longest_subsequence_count():
    # processed values (to be used in program)
    ud, vd, wd, ov=[], [], [], []

    # sum values
    u_s, v_s, w_s=0, 0, 0

    # count of values
    n=0
    
    try:

        # input xlsx file 
        wb_input=openpyxl.load_workbook("input_octant_longest_subsequence.xlsx")
        input_sh=wb_input.active

        # getting number of rows
        nr=input_sh.max_row

        # n (no of values) = number of rows(nr) - header row(1)
        n=nr-1

        for i in range(2, nr+1):
            u_s+=input_sh.cell(i,2).value
            v_s+=input_sh.cell(i,3).value
            w_s+=input_sh.cell(i,4).value

        # calculating average values
        u_s/=n; v_s/=n; w_s/=n

        #reporting data upto 9 decimal places
        u_s=int(u_s*1000000000); u_s/=1000000000
        v_s=int(v_s*1000000000); v_s/=1000000000
        w_s=int(w_s*1000000000); w_s/=1000000000

        for i in range(2, nr+1):
            x=input_sh.cell(i,2).value - u_s; x=int(x*1000000000); x/=1000000000; ud.append(x)
            y=input_sh.cell(i,3).value - v_s; y=int(y*1000000000); y/=1000000000; vd.append(y)
            z=input_sh.cell(i,4).value - w_s; z=int(z*1000000000); z/=1000000000; wd.append(z)
            o=get_octant(x,y,z); ov.append(o)

        # output xlsx file
        wb_output=Workbook()
        output_sh=wb_output.active

        # printing all initial values, average values, preprocessed values
        output_sh.cell(1,1, value='Time')
        output_sh.cell(1,2, value='U')
        output_sh.cell(1,3, value='V')
        output_sh.cell(1,4, value='W')
        for i in range(2, nr+1):
            output_sh.cell(i,1, value=input_sh.cell(i,1).value)
            output_sh.cell(i,2, value=input_sh.cell(i,2).value)
            output_sh.cell(i,3, value=input_sh.cell(i,3).value)
            output_sh.cell(i,4, value=input_sh.cell(i,4).value)


        output_sh.cell(1,5, value='U Avg'); output_sh.cell(2,5, value=u_s)
        output_sh.cell(1,6, value='V Avg'); output_sh.cell(2,6, value=v_s)
        output_sh.cell(1,7, value='W Avg'); output_sh.cell(2,7, value=w_s)

        output_sh.cell(1,8, value="U'=U - U avg")
        output_sh.cell(1,9, value="V'=V - V avg")
        output_sh.cell(1,10, value="W'=W - W avg")
        output_sh.cell(1,11, value="Octact")
        j=0
        for i in range(2,nr+1):
            output_sh.cell(i,8, value=ud[j])
            output_sh.cell(i,9, value=vd[j])
            output_sh.cell(i,10, value=wd[j])
            output_sh.cell(i,11, value=ov[j])
            j+=1
        
        # longest subsequence and count
        lsc={1:[0,0], -1:[0,0], 2:[0,0], -2:[0,0], 3:[0,0], -3:[0,0], 4:[0,0], -4:[0,0]}

        # temp octant value list, last element as terminator
        tov=ov.copy(); tov.append(0)

        # cuurent octant, current octant length
        co=ov[0]; col=1
        lsc[co][0]=1; lsc[co][1]=1


        for i in range(1,n+1):
            if(tov[i]==co):
                col+=1
            else:
                if(col>lsc[co][0]):
                    lsc[co][0]=col
                    lsc[co][1]=1
                elif(col==lsc[co][0]):
                    lsc[co][1]+=1
                co=tov[i]
                col=1

        output_sh.cell(1,13, value='Count')
        output_sh.cell(1,14, value='Longest Subsequence Length')
        output_sh.cell(1,15, value='Count')
        output_sh.cell(2,13, value=1); output_sh.cell(3,13, value=-1)
        output_sh.cell(4,13, value=2); output_sh.cell(5,13, value=-2)
        output_sh.cell(6,13, value=3); output_sh.cell(7,13, value=-3)
        output_sh.cell(8,13, value=4); output_sh.cell(9,13, value=-4)

        output_sh.cell(2,14, value=lsc[1][0]);  output_sh.cell(2,15, value=lsc[1][1])
        output_sh.cell(3,14, value=lsc[-1][0]); output_sh.cell(3,15, value=lsc[-1][1])
        output_sh.cell(4,14, value=lsc[2][0]);  output_sh.cell(4,15, value=lsc[2][1])
        output_sh.cell(5,14, value=lsc[-2][0]); output_sh.cell(5,15, value=lsc[-2][1])
        output_sh.cell(6,14, value=lsc[3][0]);  output_sh.cell(6,15, value=lsc[3][1])
        output_sh.cell(7,14, value=lsc[-3][0]); output_sh.cell(7,15, value=lsc[-3][1])
        output_sh.cell(8,14, value=lsc[4][0]); output_sh.cell(8,15, value=lsc[4][1])
        output_sh.cell(9,14, value=lsc[-4][0]); output_sh.cell(9,15, value=lsc[-4][1])

        # saving file
        wb_output.save('output_octant_longest_subsequence.xlsx')

        wb_input.close()
        wb_output.close()


    except:
        print('File Not Found')


octant_longest_subsequence_count()

end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))

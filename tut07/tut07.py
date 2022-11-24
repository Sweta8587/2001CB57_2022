
# from datetime import datetime
# start_time = datetime.now()

# #Help
# def octant_analysis(mod=5000):
# 	pass

# ##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
# ##Save all the excel files in a the output/ folder. Only xlsx to be allowed
# ## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename. 

# ###Code

# from platform import python_version
# ver = python_version()

# if ver == "3.8.10":
# 	print("Correct Version Installed")
# else:
# 	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


# mod=5000
# octant_analysis(mod)






# #This shall be the last lines of the code.
# end_time = datetime.now()
# print('Duration of Program Execution: {}'.format(end_time - start_time))

from asyncore import read                                                                                                                                # Importing libraries
from datetime import datetime
from tkinter import SOLID
import openpyxl
from openpyxl import workbook,load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from itertools import repeat
from openpyxl.styles.borders import Border, Side
 
from pandas import read_excel
from pyparsing import col
start_time = datetime.now()
import os
 
 
thin_border = Border(left=Side(style='thin'),                   # Defining border
                                                                    right=Side(style='thin'), 
                                                                    top=Side(style='thin'), 
                                                                    bottom=Side(style='thin'))
 
 
 
def find_octant(a,b,c):                                         # Function to find the octant 
              if(a>0 and b>0 and c>0):    
                           return 1
              elif(a>0 and b>0 and c<0):
                           return -1
              elif(a<0 and b>0 and c>0):
                           return 2
              elif(a<0 and b>0 and c<0):
                           return -2
              elif(a<0 and b<0 and c>0):
                           return 3
              elif(a<0 and b<0 and c<0):
                           return -3
              elif(a>0 and b<0 and c>0):
                           return 4
              elif(a>0 and b<0 and c<0):
                           return -4
 
 
def octant_analysis(mod=5000):
 
              def process_file(s,mod):                                                                                                                                     # Function to process individual file
                           in_file='input/'+s                                                                                                                                                   # Path to file
                           df=read_excel(in_file)                                                                                                                            # Reading file in dataframe
                           wb=Workbook()                                                                                                                                                                  # Starting a new workbook
                           myworkbook=wb.active                                                                                                                                                                               
                           lst=['T','U','V','W','U Avg','V Avg','W Avg',r"U'=U-U avg",r"V'=V-V avg",r"W'=W-W avg",'Octant']
                           for i in range(11):                                                                                                                                                  # Writing header of file 
                                         myworkbook.cell(row=2,column=i+1).value=lst[i]     
                           octant=[]
                           u_avg=df['U'].mean()                                        # Finding average of u,v and w                 
                           v_avg=df['V'].mean()
                           w_avg=df['W'].mean()
                           
                           myworkbook.cell(row=1,column=14).value='Overall Octant Count'
                           myworkbook.cell(row=1,column=45).value='Longest Subsequence Length'
                           myworkbook.cell(row=1,column=49).value='Longest Subsquence Length with Range'
 
                           myworkbook.cell(row=3,column=5).value=round(u_avg,3)
                           myworkbook.cell(row=3,column=6).value=round(v_avg,3)
                           myworkbook.cell(row=3,column=7).value=round(w_avg,3)
 
                           for i in df.index:
                                         myworkbook.cell(row=i+3,column=1).value=df['T'][i]
                                         myworkbook.cell(row=i+3,column=2).value=df['U'][i]
                                         myworkbook.cell(row=i+3,column=3).value=df['V'][i]
                                         myworkbook.cell(row=i+3,column=4).value=df['W'][i]
                                         myworkbook.cell(row=i+3,column=8).value=round(df['U'][i]-u_avg,3)
                                         myworkbook.cell(row=i+3,column=9).value=round(df['V'][i]-v_avg,3)
                                         myworkbook.cell(row=i+3,column=10).value=round(df['W'][i]-w_avg,3)
                                         myworkbook.cell(row=i+3,column=11).value=find_octant(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg)
                                         octant.append(find_octant(df['U'][i]-u_avg,df['V'][i]-v_avg,df['W'][i]-w_avg))
                           
                           def octant_range_names(mod=5000):
                                         octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
                                         dic={}                                                          # creating dictionary for mapping 
                                         my_dic={}                                                      # Creating dictionary with opposite key value pair than 'dic'
                                         
                                         for i in range(0,4):                                            # dic[1]=0,dic[-1]=-1,...
                                                       dic[i+1]=2*i+1-1                                            # my_dic[0]=1,my_dic[1]=-1,...
                                                       dic[-(i+1)]=2*(i+1)-1
                                                       my_dic[2*i+1-1]=i+1
                                                       my_dic[2*(i+1)-1]=-(i+1)
 
                                         def find_rank_of_list(lst):                                     # Function to find the rank list from count values of all octants
                                                       temp_lst=lst.copy()
                                                       temp_lst.sort(reverse=True)
                                                       res=[]
 
                                                       for i in lst:
                                                                    for j in range(0,8):
                                                                                  if(i==temp_lst[j]):
                                                                                                res.append(j+1)
                                                                                                break
                                                       return res                                                  # Returning the ranked list
                                         
                                         def find_1st_rank(lst):                                         # Finding the octant which has rank 1 in the given rank list
                                                       for i in range(8):
                                                                    if(lst[i]==1):
                                                                                  return my_dic[i]
 
                                         def count_rank1(lst,x):                                         # Finding the count of rank 1 in the rank 1 mod values of octant x
                                                       sum=0
                                                       for i in lst:
                                                                    if(x==i):
                                                                                  sum+=1
                                                       return sum                                                  # Return the count
                                         
                                         my_matr=[]                                                  # Matrix to store rank list for different mod values
                                         rank1_list=[]                                                   # List to store the octants which have rank 1 in different mod ranges and overall
                                         myworkbook=wb.active
                                         myworkbook['M4']='Mod '+str(mod)                                           # Putting the string 'User Input' at its specified place
 
                                         matrix=[]                                                       # 2-d matrix for storing octants within ranges
                                         count=[0]*9                                                     # Creating a list for storing elements of 9 columns
 
                                         count[0]='Octant ID'                                            # Storing header list in 'count' list
 
                                         for i in range(0,4):
                                                       count[2*i+1]=(i+1)
                                                       count[2*(i+1)]=-(i+1)
                                         matrix.append(count)                                            # Appending header list in matrix
                                         for i in range(13,22):                                          # Writing header list in worksheet
                                                       myworkbook.cell(row=3,column=i+1).value=count[i-13]
                                                       myworkbook.cell(row=3,column=i+1).border=thin_border
                                                       if(i>13):
                                                                    myworkbook.cell(row=3,column=i+9).value='Rank Octant '+str(count[i-13])
                                                                    myworkbook.cell(row=3,column=i+9).border=thin_border
                                         myworkbook.cell(row=3,column=31).value='Rank1 Octant ID'
                                         myworkbook.cell(row=3,column=32).value='Rank1 Octant Name'
                                         myworkbook.cell(row=3,column=31).border=thin_border
                                         myworkbook.cell(row=3,column=32).border=thin_border
                                         count=[0]*9                                                     # Resetting values in list 'count'
                           
                                         for i in octant:                                                # Finding total count of values in different octants
                                                       if(i==1):
                                                                    count[1]=count[1]+1
                                                       elif(i==-1):
                                                                    count[2]=count[2]+1
                                                       elif(i==2):
                                                                    count[3]=count[3]+1
                                                       elif(i==-2):
                                                                    count[4]=count[4]+1
                                                       elif(i==3):
                                                                    count[5]=count[5]+1
                                                       elif(i==-3):
                                                                    count[6]=count[6]+1
                                                       elif(i==4):
                                                                    count[7]=count[7]+1
                                                       elif(i==-4):
                                                                    count[8]=count[8]+1
                                         yellow = "00FFFF00"
                                         count[0]='Overall Count'                                        # Creating overall count row
                                         matrix.append(count)                                           
                                         for i in range(13,22):                                          # Writing overall count in worksheet
                                                       myworkbook.cell(row=4,column=i+1).value=count[i-13]
                                                       myworkbook.cell(row=4,column=i+1).border=thin_border
                                         count.pop(0)                                                    # Removing the header from list
                                         rank=find_rank_of_list(count)                                   # Find the rank list 
                                         rank1_list.append(find_1st_rank(rank))                          # Finding the rank 1 octant and appending in rank1_list
                                         my_matr.append(rank)                                        # Appending rank list in the matrix
                                         for i in range(8):                                              # Writing overall count in worksheet
                                                       myworkbook.cell(row=4,column=23+i).value=my_matr[0][i]
                                                       myworkbook.cell(row=4,column=23+i).border=thin_border
                                                       if(my_matr[0][i]==1):
                                                                     myworkbook.cell(row=4,column=23+i).fill=PatternFill(start_color=yellow,end_color=yellow,fill_type="solid")
                                         myworkbook.cell(row=4,column=31).value=rank1_list[0]
                                         myworkbook.cell(row=4,column=32).value=octant_name_id_mapping[str(rank1_list[0])]
                                         myworkbook.cell(row=4,column=31).border=thin_border
                                         myworkbook.cell(row=4,column=32).border=thin_border


# defining mod value
mod=5000

# importing openpyxl and nan and loading workbook
try:
    from cmath import nan
    import openpyxl

    wb = openpyxl.load_workbook(r'C:\Users\DELL\OneDrive\Desktop\temp_octant\input_octant_transition_identify - Copy.xlsx')
except:
    print("there is error in loading workbook check your file directory and import openpyxl")
    exit()


    #### calculating average value
try:
    sheet = wb.active
    Uavg=0
    Vavg=0
    Wavg=0
    ls=[]
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=29746, max_col=4):
        lst1=[]
        for cell in row:
            lst1.append(cell.value)
        ls.append(lst1)
    for i in range(1,29746):
        Uavg=Uavg+ls[i][1]
        Vavg=Vavg+ls[i][2]
        Wavg=Wavg+ls[i][3]
    Uavg=Uavg/29745
    Vavg=Vavg/29745
    Wavg=Wavg/29745

    lst_avg=[["Uavg","Vavg","Wavg"],[Uavg,Vavg,Wavg]]
except:
    print("there is error in calculating average value")
    exit()

# writing nan in tansition value

try:
    i=0
    for row in sheet.iter_rows(min_row=1, min_col=12, max_row=29777, max_col=21):
        j=0
        for cell in row:
            cell.value=nan
            j=j+1
        i=i+1
except:
    print("there is some error in transition count")    


#creating octant of given value

try:
    i=0
    for row in sheet.iter_rows(min_row=1, min_col=5, max_row=2, max_col=7):
        j=0
        for cell in row:
            cell.value=lst_avg[i][j]
            j=j+1
        i=i+1
    lst_newval=[]
    for i in range(1,29746):
        lst_newval.append([ls[i][1]-Uavg,ls[i][2]-Vavg,ls[i][3]-Wavg])
    sheet.cell(row=1,column=8).value="Uavg'"
    sheet.cell(row=1,column=9).value="Vavg'"
    sheet.cell(row=1,column=10).value="Wavg'"    
    sheet.cell(row=1,column=11).value="Octant"    

    i=0
    for row in sheet.iter_rows(min_row=2, min_col=8, max_row=29746, max_col=10):
        j=0
        for cell in row:
            cell.value=lst_newval[i][j]
            j=j+1
        i=i+1

    lst_octant = []
    for p in lst_newval:
        if(p[0]>=0):
            if(p[1]>=0):
                if(p[2]>=0):
                    lst_octant.append(1)
                else:
                    lst_octant.append(-1)
            else:
                if(p[2]>=0):
                    lst_octant.append(4)
                else:
                    lst_octant.append(-4)
        else:
            if(p[1]>=0):
                if(p[2]>=0):
                    lst_octant.append(2)
                else:
                    lst_octant.append(-2)
            else:
                if(p[2]>=0):
                    lst_octant.append(3)
                else:
                    lst_octant.append(-3)

    i=0
    for row in sheet.iter_rows(min_row=2, min_col=11, max_row=29746, max_col=11):
        for cell in row:
            cell.value=lst_octant[i]
        i=i+1
except:
    print("there is error in creating octant value")
    exit()

# there may be some error in creating overall count

try:
    t=0
    if(29746%mod==0):
        t=29746//mod
    else:
        t=(29746//mod)+1

    sheet.cell(row=1,column=14).value="+1"
    sheet.cell(row=1,column=15).value="-1"
    sheet.cell(row=1,column=16).value="+2"
    sheet.cell(row=1,column=17).value="-2"
    sheet.cell(row=1,column=18).value="+3"
    sheet.cell(row=1,column=19).value="-3"
    sheet.cell(row=1,column=20).value="+4"
    sheet.cell(row=1,column=21).value="-4"
    sheet.cell(row=2,column=13).value="Overall Count"
    sheet.cell(row=3,column=12).value="User Input"
    tt=str(mod)
    sheet.cell(row=3,column=13).value="mod"+" "+tt
    lst_overall_count = [0,0,0,0,0,0,0,0]
    for valu in lst_octant:
        if valu==1:
            lst_overall_count[0]=lst_overall_count[0]+1
        if valu==-1:
            lst_overall_count[1]=lst_overall_count[1]+1
        if valu==2:
            lst_overall_count[2]=lst_overall_count[2]+1
        if valu==-2:
            lst_overall_count[3]=lst_overall_count[3]+1
        if valu==3:
            lst_overall_count[4]=lst_overall_count[4]+1
        if valu==-3:
            lst_overall_count[5]=lst_overall_count[5]+1
        if valu==4:
            lst_overall_count[6]=lst_overall_count[6]+1
        if valu==-4:
            lst_overall_count[7]=lst_overall_count[7]+1
    for row in sheet.iter_rows(min_row=2, min_col=14, max_row=2, max_col=21):
        j=0
        for cell in row:
            cell.value=lst_overall_count[j]
            j=j+1
    lst_hh=[]
    for j in range(t):
        if(j==t-1):
            ttm1=str(mod*j)
            ttm=ttm1+"-"+"29745"
            lst_hh.append(ttm)
        else:
            if(j==0):
                ttm1=".0000"
            else:
                ttm1=str(mod*j) 
            ttm2=str(mod*(j+1)-1)
            ttm=ttm1+"-"+ttm2
            lst_hh.append(ttm) 
    lst_hh.append("Verified")
except:
    print("there is error in creating overall count value and verifing it")
    exit()


# updating overall count and transition
try:
    lst_hh_val=[]
    for j in range(t):
        lst_hh_temp=[0,0,0,0,0,0,0,0]
        if(j==t-1):
            y=29745
        else:
            y=mod*(j+1)
        for valu in range(mod*j,y):
            if lst_octant[valu]==1:
                lst_hh_temp[0]=lst_hh_temp[0]+1
            if lst_octant[valu]==-1:
                lst_hh_temp[1]=lst_hh_temp[1]+1
            if lst_octant[valu]==2:
                lst_hh_temp[2]=lst_hh_temp[2]+1
            if lst_octant[valu]==-2:
                lst_hh_temp[3]=lst_hh_temp[3]+1
            if lst_octant[valu]==3:
                lst_hh_temp[4]=lst_hh_temp[4]+1
            if lst_octant[valu]==-3:
                lst_hh_temp[5]=lst_hh_temp[5]+1
            if lst_octant[valu]==4:
                lst_hh_temp[6]=lst_hh_temp[6]+1
            if lst_octant[valu]==-4:
                lst_hh_temp[7]=lst_hh_temp[7]+1


        lst_hh_val.append(lst_hh_temp)

    lst_verified=[0,0,0,0,0,0,0,0]
    for i in range(t):
        for j in range(8):
            lst_verified[j]=lst_verified[j]+lst_hh_val[i][j]

    j=0
    for row in sheet.iter_rows(min_row=4, min_col=13, max_row=t+4, max_col=13):
        for cell in row:
            cell.value=lst_hh[j]
        j=j+1

    i=0
    for row in sheet.iter_rows(min_row=4, min_col=14, max_row=t+3, max_col=21):
        j=0
        for cell in row:
            cell.value=lst_hh_val[i][j]
            j=j+1
        i=i+1

    for row in sheet.iter_rows(min_row=t+4, min_col=14, max_row=t+4, max_col=21):
        j=0
        for cell in row:
            cell.value=lst_verified[j]
            j=j+1
except:
    print("there is error in updating in verified count")
   